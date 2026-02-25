#!/usr/bin/env python3
"""
DPA Contract Review Tool

Fetches an incoming DPA from a Google Doc (paragraph by paragraph),
compares it against a ClearTax playbook and a rulebook loaded from xlsx,
flags deviations by adding comments directly on the Google Doc,
and writes a structured flags.json.

Usage:
    python main.py <input_doc_url_or_id> [--playbook <playbook_url_or_id>]

The rulebook is always loaded from "DPA Rulebook.xlsx" in the project dir.
Google credentials come from "credentials.json" in the project dir.
"""

import json
import math
import os
import sys
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import numpy as np
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

# ---------------------------------------------------------------------------
# Load .env if present
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
_env_path = BASE_DIR / ".env"
if _env_path.exists():
    for _line in _env_path.read_text().splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _, _v = _line.partition("=")
            os.environ.setdefault(_k.strip(), _v.strip().strip("\"'"))

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
RULEBOOK_PATH = BASE_DIR / "DPA Rulebook.xlsx"
PLAYBOOK_PATH = BASE_DIR / "Clear Tax_DPA.docx"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_PATH = OUTPUT_DIR / "flags.json"

EMBED_MODEL = os.environ.get("EMBED_MODEL", "BAAI/bge-base-en-v1.5")
EMBED_MODEL_FALLBACK = "all-MiniLM-L6-v2"
MATCH_THRESHOLD_STRONG = 0.70
MATCH_THRESHOLD_PARTIAL = 0.45
RULE_MATCH_THRESHOLD = 0.58
KEYWORD_BOOST = 0.08

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
LLM_MODEL = "claude-sonnet-4-20250514"

CREDS_PATH = BASE_DIR / "credentials.json"

# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass
class Clause:
    id: str
    text: str
    section: str = ""
    source: str = ""       # "input" or "playbook"
    start_index: int = 0   # Google Docs character offset (for anchoring)
    end_index: int = 0
    raw_text: str = ""     # Original paragraph text from Google Doc (for comment anchoring)


@dataclass
class Rule:
    rule_id: str
    source: str       # "legal" or "infosec"
    clause: str
    subclause: str
    risk: str
    response: str


# ---------------------------------------------------------------------------
# 1. GOOGLE AUTH
# ---------------------------------------------------------------------------

_google_creds = None


def get_google_creds():
    """Load Google service-account credentials from credentials.json."""
    global _google_creds
    if _google_creds is not None:
        return _google_creds

    from google.oauth2 import service_account

    if not CREDS_PATH.exists():
        print(f"Error: {CREDS_PATH} not found.")
        print("Place your Google service-account key as credentials.json")
        sys.exit(1)

    _google_creds = service_account.Credentials.from_service_account_file(
        str(CREDS_PATH),
        scopes=[
            "https://www.googleapis.com/auth/documents",
            "https://www.googleapis.com/auth/drive",
        ],
    )
    return _google_creds


def extract_doc_id(url_or_id: str) -> str:
    """Extract Google Doc ID from a URL or return as-is if already an ID."""
    # Full URL: https://docs.google.com/document/d/DOC_ID/...
    m = re.search(r"/document/d/([a-zA-Z0-9_-]+)", url_or_id)
    if m:
        return m.group(1)
    # Bare ID (alphanumeric, 20+ chars)
    if re.match(r"^[a-zA-Z0-9_-]{20,}$", url_or_id):
        return url_or_id
    print(f"Error: Cannot parse Google Doc ID from: {url_or_id}")
    sys.exit(1)


# ---------------------------------------------------------------------------
# 2. LOAD RULEBOOK FROM XLSX (dynamic, no hardcoding)
# ---------------------------------------------------------------------------

def load_rulebook(path: Path) -> list[Rule]:
    """
    Parse DPA Rulebook.xlsx dynamically.
    Reads Legal and Infosec sheets, extracts rules with clause/subclause/risk/response.
    Handles hierarchical rows where empty Clause means continuation of previous.
    """
    import openpyxl

    if not path.exists():
        print(f"Error: Rulebook not found: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(path), read_only=True)
    rules: list[Rule] = []
    rule_idx = 0

    for sheet_name in wb.sheetnames:
        name_lower = sheet_name.lower()
        if "legal" in name_lower:
            source = "legal"
        elif "infosec" in name_lower:
            source = "infosec"
        else:
            continue  # skip non-rule sheets (General automation, etc.)

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # --- Discover column indices from header row ---
        header = [str(c).lower().strip() if c else "" for c in rows[0]]

        def find_col(*keywords):
            for i, h in enumerate(header):
                if any(kw in h for kw in keywords):
                    return i
            return None

        col_clause = find_col("clause")
        col_sub = find_col("sub-clause", "subclause", "sub clause")
        col_risk = find_col("risk")
        col_resp = find_col("response")

        if col_clause is None and col_sub is None:
            continue  # can't parse this sheet

        # If no separate subclause column, use clause column for both
        if col_sub is None:
            col_sub = col_clause

        # --- Parse data rows ---
        current_clause = ""
        for row in rows[1:]:
            def cell(idx):
                if idx is None or idx >= len(row):
                    return ""
                return str(row[idx]).strip() if row[idx] else ""

            clause_val = cell(col_clause)
            sub_val = cell(col_sub)
            risk_val = cell(col_risk)
            resp_val = cell(col_resp)

            # Track parent clause for hierarchical rows
            if clause_val:
                current_clause = clause_val

            # Skip empty / section-header rows (no risk or no response)
            if not risk_val or risk_val.lower() in ("none", "risk", ""):
                continue
            if not resp_val:
                continue

            rule_idx += 1
            rules.append(Rule(
                rule_id=f"{source}_{rule_idx}",
                source=source,
                clause=current_clause,
                subclause=sub_val or current_clause,
                risk=risk_val,
                response=resp_val,
            ))

    wb.close()
    return rules


# ---------------------------------------------------------------------------
# 3. FETCH PARAGRAPHS (Google Docs or .docx)
# ---------------------------------------------------------------------------

def fetch_gdoc_paragraphs(doc_id: str) -> list[dict]:
    """
    Fetch paragraphs from a Google Doc via the Docs API.
    Returns list of {"text": ..., "start_index": ..., "end_index": ...}.
    """
    from googleapiclient.discovery import build

    creds = get_google_creds()
    service = build("docs", "v1", credentials=creds, cache_discovery=False)
    doc = service.documents().get(documentId=doc_id).execute()

    paragraphs: list[dict] = []
    for element in doc.get("body", {}).get("content", []):
        para = element.get("paragraph")
        if not para:
            continue
        start = element.get("startIndex", 0)
        end = element.get("endIndex", 0)
        parts = []
        for elem in para.get("elements", []):
            tr = elem.get("textRun")
            if tr:
                parts.append(tr.get("content", ""))
        full = "".join(parts).strip()
        if full:
            paragraphs.append({
                "text": full,
                "start_index": start,
                "end_index": end,
            })
    return paragraphs


def fetch_docx_paragraphs(path: Path) -> list[dict]:
    """Read paragraphs from a local .docx file."""
    from docx import Document

    doc = Document(str(path))
    paragraphs: list[dict] = []
    offset = 0
    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            paragraphs.append({
                "text": text,
                "start_index": offset,
                "end_index": offset + len(text),
            })
            offset += len(text) + 1
    return paragraphs


# ---------------------------------------------------------------------------
# 4. CLAUSE EXTRACTION (content-based, no hardcoded styles)
# ---------------------------------------------------------------------------

# Words that appear in clauses but not in section headers
_CLAUSE_VERBS = {
    "shall", "will", "must", "may", "agrees", "acknowledges",
    "warrants", "represents", "confirms", "ensures", "undertakes",
}

_DEF_RE = re.compile(
    r'^\s*[\"\u201c\u2018][^\"\u201d\u2019]+[\"\u201d\u2019]\s*'
    r'(,\s*[\"\u201c\u2018][^\"\u201d\u2019]+[\"\u201d\u2019]\s*)*'
    r'(means|shall\s+mean|shall\s+have)',
    re.IGNORECASE,
)

_SKIP_RE = re.compile(
    r"(^IN WITNESS WHEREOF|^WHEREAS\b|^NOW,?\s*THEREFORE|"
    r"^Sign\s*:|^Signed?\s*:|_{5,})",
    re.IGNORECASE,
)

_APPENDIX_RE = re.compile(
    r"^(APPENDIX|ANNEX|ATTACHMENT)\s", re.IGNORECASE,
)


def _is_definition(text: str) -> bool:
    low = text.lower()
    if "shall mean" in low or "shall have the meaning" in low:
        return True
    return bool(_DEF_RE.match(text))


def _is_section_header(text: str) -> tuple[bool, str, str]:
    """Returns (is_header, section_name, remainder_after_header)."""
    stripped = text.strip()
    # ALL-CAPS short line
    if len(stripped) < 100 and stripped.upper() == stripped and stripped.isascii():
        if any(c.isalpha() for c in stripped):
            return True, stripped.rstrip(".:;, "), ""
    # Title-case label followed by period: "Subprocessing. <optional remainder>"
    m = re.match(
        r'^(\d{1,3}\.?\s+)?([A-Z][A-Za-z,;/&\s\-\(\)\']+?)\.\s*(.*)',
        stripped, re.DOTALL,
    )
    if m:
        label = m.group(2).strip()
        remainder = m.group(3).strip()
        if len(label) > 100:
            return False, "", ""
        if _CLAUSE_VERBS & set(label.lower().split()):
            return False, "", ""
        if len(label.split()) <= 15:
            return True, label, remainder
    return False, "", ""


def _is_appendix_boundary(text: str) -> bool:
    stripped = text.strip()
    up = stripped.upper()
    if up in ("EXHIBIT A", "EXHIBIT B", "EXHIBIT C"):
        return False
    if _APPENDIX_RE.match(stripped):
        return True
    if re.match(
        r"^(STANDARD CONTRACTUAL|LIST OF SUB-?PROCESSORS|"
        r"TECHNICAL AND ORGANI[SZ]ATIONAL|DESCRIPTION OF TECHNICAL)",
        stripped, re.IGNORECASE,
    ):
        return True
    return False


def _is_preamble(text: str) -> bool:
    low = text.lower()
    return any(p in low for p in [
        "this data processing agreement",
        "the dpa shall form",
        "entering into this dpa",
        "in the event of inconsistencies",
        "the parties have agreed",
        "order of priority shall be",
        "hereinafter referred to",
        "referred to individually as",
        "seek to implement a data processing",
        "wish to lay down their rights",
        "in consideration of the mutual covenants",
    ])


def extract_clauses(paragraphs: list[dict], source: str) -> list[Clause]:
    """
    Extract substantive clauses from a list of paragraphs.
    Content-based filtering — works with any DPA regardless of formatting.
    """
    clauses: list[Clause] = []
    section = ""
    in_appendix = False
    in_definitions = False
    idx = 0
    i = 0

    while i < len(paragraphs):
        p = paragraphs[i]
        text = p["text"]

        # Stop at appendices / annexes
        if _is_appendix_boundary(text):
            in_appendix = True
            i += 1
            continue
        if in_appendix:
            i += 1
            continue

        # Skip boilerplate
        if _SKIP_RE.search(text):
            i += 1
            continue

        # Section headers
        is_hdr, sec_name, remainder = _is_section_header(text)
        if is_hdr:
            section = sec_name
            in_definitions = any(
                kw in sec_name.lower()
                for kw in ("definition", "defined term", "interpretation")
            )
            if len(remainder) >= 40:
                text = remainder
            else:
                i += 1
                continue

        # Skip definitions
        if in_definitions:
            i += 1
            continue
        if _is_definition(text):
            i += 1
            continue

        # Skip preamble
        if _is_preamble(text):
            i += 1
            continue

        # Skip very short lines
        if len(text) < 35:
            i += 1
            continue

        # Merge short continuations (but not across headers)
        merged_text = text
        merged_end = p["end_index"]
        while i + 1 < len(paragraphs):
            nxt = paragraphs[i + 1]["text"]
            if len(nxt) >= 40:
                break
            if _SKIP_RE.search(nxt) or _is_appendix_boundary(nxt) or _is_definition(nxt):
                break
            nxt_hdr, _, _ = _is_section_header(nxt)
            if nxt_hdr:
                break
            merged_text += " " + nxt
            merged_end = paragraphs[i + 1]["end_index"]
            i += 1

        if len(merged_text) >= 40:
            idx += 1
            clauses.append(Clause(
                id=f"{source}_{idx}",
                text=merged_text,
                section=section,
                source=source,
                start_index=p["start_index"],
                end_index=merged_end,
                raw_text=p["text"],  # original first paragraph (unprocessed)
            ))

        i += 1

    return clauses


# ---------------------------------------------------------------------------
# 5. EMBEDDING & MATCHING
# ---------------------------------------------------------------------------

_model: Optional[SentenceTransformer] = None


def get_model() -> SentenceTransformer:
    global _model
    if _model is None:
        try:
            print(f"  Loading embedding model: {EMBED_MODEL}...")
            _model = SentenceTransformer(EMBED_MODEL)
        except Exception as e:
            print(f"  Could not load {EMBED_MODEL}: {e}")
            print(f"  Falling back to {EMBED_MODEL_FALLBACK}...")
            _model = SentenceTransformer(EMBED_MODEL_FALLBACK)
    return _model


def embed_texts(texts: list[str]) -> np.ndarray:
    return get_model().encode(texts, show_progress_bar=False, convert_to_numpy=True)


def match_clauses(
    input_clauses: list[Clause],
    playbook_clauses: list[Clause],
) -> list[tuple[Clause, Clause, float, str]]:
    """For each input clause, find the best-matching playbook clause."""
    print("  Computing embeddings...")
    inp_emb = embed_texts([c.text for c in input_clauses])
    pb_emb = embed_texts([c.text for c in playbook_clauses])
    sim_matrix = cosine_similarity(inp_emb, pb_emb)

    results = []
    for i, ic in enumerate(input_clauses):
        j = int(np.argmax(sim_matrix[i]))
        score = float(sim_matrix[i][j])
        if score >= MATCH_THRESHOLD_STRONG:
            mt = "strong"
        elif score >= MATCH_THRESHOLD_PARTIAL:
            mt = "partial"
        else:
            mt = "new_clause"
        results.append((ic, playbook_clauses[j], score, mt))
    return results


def match_rules(clause: Clause, rules: list[Rule]) -> list[tuple[Rule, float]]:
    """Find applicable rules for a clause (semantic similarity + keyword boost)."""
    clause_emb = embed_texts([clause.text])
    rule_texts = [f"{r.clause}: {r.subclause}" for r in rules]
    rule_emb = embed_texts(rule_texts)
    sims = cosine_similarity(clause_emb, rule_emb)[0]

    clause_low = clause.text.lower()
    matched: list[tuple[Rule, float]] = []
    for j, rule in enumerate(rules):
        score = float(sims[j])
        # Keyword boost: if words from the rule clause appear in the input text
        rule_keywords = set(re.findall(r"[a-z]{4,}", rule.clause.lower()))
        clause_words = set(re.findall(r"[a-z]{4,}", clause_low))
        overlap = rule_keywords & clause_words
        if len(overlap) >= 2:
            score = min(score + KEYWORD_BOOST, 1.0)
        if score >= RULE_MATCH_THRESHOLD:
            matched.append((rule, score))

    matched.sort(key=lambda x: -x[1])
    return matched[:5]


def apply_rule_specificity(
    all_clause_rules: list[list[tuple[Rule, float]]],
) -> list[list[tuple[Rule, float]]]:
    """
    Penalize broad rules that match too many clauses (rule specificity scoring).
    Broad rules like 'data protection laws' match everything — this reduces their weight.
    specificity = 1.0 / log(1 + num_clauses_matched)
    """
    # Count how many clauses each rule triggered on
    rule_trigger_count: dict[str, int] = {}
    for clause_rules in all_clause_rules:
        for rule, _ in clause_rules:
            rule_trigger_count[rule.rule_id] = rule_trigger_count.get(rule.rule_id, 0) + 1

    # Apply specificity penalty and re-filter
    adjusted: list[list[tuple[Rule, float]]] = []
    for clause_rules in all_clause_rules:
        new_rules: list[tuple[Rule, float]] = []
        for rule, score in clause_rules:
            count = rule_trigger_count[rule.rule_id]
            specificity = 1.0 / math.log(1 + count)
            adjusted_score = score * specificity
            if adjusted_score >= RULE_MATCH_THRESHOLD:
                new_rules.append((rule, adjusted_score))
        new_rules.sort(key=lambda x: -x[1])
        adjusted.append(new_rules[:5])
    return adjusted


# ---------------------------------------------------------------------------
# 6. LLM ANALYSIS
# ---------------------------------------------------------------------------

def _build_prompt(
    inp: Clause, pb: Clause, sim: float, mt: str,
    rules: list[tuple[Rule, float]],
) -> str:
    pb_block = (
        f'MATCHED PLAYBOOK CLAUSE (ClearTax Standard):\n'
        f'"{pb.text}"\nSimilarity: {sim:.2f} ({mt} match)'
        if mt != "new_clause" else
        f'NO MATCHING PLAYBOOK CLAUSE (similarity: {sim:.2f})\n'
        f'This is a NEW obligation not in ClearTax\'s standard DPA.'
    )
    rules_block = ""
    if rules:
        rules_block = "\n\nAPPLICABLE RULES FROM INTERNAL RULEBOOK:\n"
        for r, _ in rules:
            rules_block += (
                f"\n- [{r.rule_id}] {r.clause} (Risk: {r.risk})\n"
                f"  Condition: {r.subclause}\n"
                f"  Required: {r.response}\n"
            )

    return f"""You are a legal analyst reviewing a DPA for ClearTax (Defmacro Software Pvt Ltd), the data processor.
The incoming DPA is from a customer (the data controller).

Analyze this clause and determine compliance with ClearTax's playbook and internal rules.

INCOMING CLAUSE:
"{inp.text}"

{pb_block}{rules_block}

Classify as: "compliant" / "deviation_minor" / "deviation_major" / "non_compliant"
Risk: "High" / "Medium" / "Low"
Confidence: a float 0.0–1.0 indicating how confident you are in this assessment.

Respond ONLY with this JSON (no markdown fences):
{{"classification": "...", "risk_level": "...", "explanation": "...", "suggested_redline": "...", "confidence": 0.0}}"""


def _call_llm(prompt: str) -> dict:
    import anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    resp = client.messages.create(
        model=LLM_MODEL, max_tokens=1024,
        messages=[{"role": "user", "content": prompt}],
    )
    text = resp.content[0].text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1].rsplit("```", 1)[0].strip()
    return json.loads(text)


def _compute_confidence(sim: float, rules: list[tuple[Rule, float]]) -> float:
    """Compute confidence score (0.0–1.0) for heuristic analysis."""
    deviation_signal = (1 - sim) * 0.4
    avg_rule_score = sum(s for _, s in rules) / len(rules) if rules else 0.0
    rule_score_signal = avg_rule_score * 0.4
    rule_count_signal = min(len(rules) / 3, 1.0) * 0.2
    return round(min(deviation_signal + rule_score_signal + rule_count_signal, 1.0), 3)


def _heuristic(
    _inp: Clause, _pb: Clause, sim: float, mt: str,
    rules: list[tuple[Rule, float]],
) -> dict:
    # Only consider rules with high match confidence (score > 0.62)
    strong_rules = [(r, s) for r, s in rules if s > 0.62]
    rule_risks = [r.risk for r, _ in strong_rules]
    max_risk = "High" if "High" in rule_risks else ("Medium" if "Medium" in rule_risks else "Low")
    rule_ids = ", ".join(r.rule_id for r, _ in strong_rules)
    top_response = strong_rules[0][0].response if strong_rules else ""
    confidence = _compute_confidence(sim, strong_rules)

    # Strong match + no meaningful rules → compliant
    if mt == "strong" and not strong_rules:
        return dict(classification="compliant", risk_level="Low",
                    explanation=f"Matches ClearTax standard (sim={sim:.2f}).",
                    suggested_redline="", confidence=confidence)

    # Partial match + no meaningful rules → likely compliant (just different wording)
    if mt == "partial" and not strong_rules:
        return dict(classification="compliant", risk_level="Low",
                    explanation=f"Partial match (sim={sim:.2f}). No policy rules triggered.",
                    suggested_redline="", confidence=confidence)

    # Strong match but triggers rules → minor deviation
    if mt == "strong":
        return dict(classification="deviation_minor", risk_level=max_risk,
                    explanation=f"Matches standard (sim={sim:.2f}) but triggers rules: {rule_ids}.",
                    suggested_redline=top_response, confidence=confidence)

    # Partial match + rules triggered → severity depends on risk
    if mt == "partial":
        cls = "deviation_major" if max_risk == "High" else "deviation_minor"
        return dict(classification=cls, risk_level=max_risk,
                    explanation=f"Partial match (sim={sim:.2f}). Triggered: {rule_ids}.",
                    suggested_redline=top_response, confidence=confidence)

    # New clause (no playbook equivalent)
    if not strong_rules:
        # New clause but no rules care about it → compliant (boilerplate addition)
        return dict(classification="compliant", risk_level="Low",
                    explanation=f"New clause (sim={sim:.2f}). No policy rules triggered.",
                    suggested_redline="", confidence=confidence)

    return dict(
        classification="deviation_major" if max_risk == "High" else "deviation_minor",
        risk_level=max_risk,
        explanation=f"New clause not in ClearTax standard (sim={sim:.2f}). Triggered: {rule_ids}.",
        suggested_redline=top_response, confidence=confidence,
    )


def analyze_clause(inp, pb, sim, mt, rules, use_llm) -> dict:
    if use_llm:
        try:
            return _call_llm(_build_prompt(inp, pb, sim, mt, rules))
        except Exception as e:
            print(f"    LLM error: {e}. Using heuristic.")
    return _heuristic(inp, pb, sim, mt, rules)


# ---------------------------------------------------------------------------
# 7. ADD COMMENTS TO GOOGLE DOC
# ---------------------------------------------------------------------------

def clear_old_comments(doc_id: str) -> int:
    """Delete all existing comments on the doc (from previous runs)."""
    from googleapiclient.discovery import build

    creds = get_google_creds()
    drive = build("drive", "v3", credentials=creds, cache_discovery=False)

    deleted = 0
    page_token = None
    while True:
        resp = drive.comments().list(
            fileId=doc_id, fields="comments(id,content),nextPageToken",
            pageToken=page_token, includeDeleted=False,
        ).execute()
        for comment in resp.get("comments", []):
            # Only delete comments that look like ours (start with risk tag)
            content = comment.get("content", "")
            if content.startswith("[High Risk]") or content.startswith("[Medium Risk]") or content.startswith("[Low Risk]"):
                try:
                    drive.comments().delete(
                        fileId=doc_id, commentId=comment["id"],
                    ).execute()
                    deleted += 1
                except Exception:
                    pass
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return deleted


def _extract_doc_plain_text(doc_id: str) -> str:
    """Fetch the full plain-text content of a Google Doc (for exact quoting)."""
    from googleapiclient.discovery import build

    creds = get_google_creds()
    docs = build("docs", "v1", credentials=creds, cache_discovery=False)
    doc = docs.documents().get(documentId=doc_id).execute()

    # Build a character-indexed plain text from the doc body
    parts: list[tuple[int, str]] = []
    for element in doc.get("body", {}).get("content", []):
        para = element.get("paragraph")
        if not para:
            continue
        for elem in para.get("elements", []):
            tr = elem.get("textRun")
            if tr:
                idx = elem.get("startIndex", 0)
                parts.append((idx, tr.get("content", "")))
    parts.sort(key=lambda x: x[0])

    # Reconstruct full text with correct offsets
    if not parts:
        return ""
    max_end = max(idx + len(txt) for idx, txt in parts)
    buf = [" "] * max_end
    for idx, txt in parts:
        for j, ch in enumerate(txt):
            if idx + j < max_end:
                buf[idx + j] = ch
    return "".join(buf)


def add_comments_to_doc(doc_id: str, flags: list[dict]) -> int:
    """
    Add comments to flagged paragraphs.
    Extracts exact text from the Google Doc at character offsets for
    quotedFileContent to maximize anchoring accuracy.
    """
    from googleapiclient.discovery import build
    import json as _json

    creds = get_google_creds()
    drive = build("drive", "v3", credentials=creds, cache_discovery=False)

    # Get the exact plain text of the doc for precise quoting
    print("    Fetching doc text for exact quoting...")
    full_text = _extract_doc_plain_text(doc_id)

    added = 0
    for flag in flags:
        if flag["classification"] == "compliant":
            continue

        # Build the comment text
        risk = flag["risk_level"]
        cls = flag["classification"].replace("_", " ").title()

        rule_lines = []
        for r in flag["triggered_rules"]:
            rule_lines.append(f"  - [{r['source'].upper()}] {r['clause']} (Risk: {r['risk']})")

        comment_text = f"[{risk} Risk] {cls}\n\n"
        comment_text += f"{flag['explanation']}\n"

        if rule_lines:
            comment_text += "\nRulebook violations:\n" + "\n".join(rule_lines) + "\n"

        if flag["suggested_redline"]:
            comment_text += f"\nSuggested redline:\n{flag['suggested_redline']}"

        # Extract EXACT text from doc at the character offsets
        start = flag.get("start_index", 0)
        end = flag.get("end_index", 0)
        exact_text = full_text[start:end].strip()
        # Use first 300 chars for the quote
        quoted = exact_text[:300] if exact_text else flag["input_text"][:300]

        # Build anchor using txt region classifier (best-effort)
        anchor = _json.dumps({
            "r": "head",
            "a": [{"txt": {"o": start, "l": end - start, "ml": len(full_text)}}],
        })

        body = {
            "content": comment_text,
            "anchor": anchor,
            "quotedFileContent": {
                "mimeType": "text/plain",
                "value": quoted,
            },
        }

        try:
            drive.comments().create(
                fileId=doc_id, body=body, fields="id,anchor",
            ).execute()
            added += 1
        except Exception as e:
            print(f"    Could not add comment for {flag['flag_id']}: {e}")

    return added


# ---------------------------------------------------------------------------
# 8. HIGHLIGHT FLAGGED PARAGRAPHS IN GOOGLE DOC
# ---------------------------------------------------------------------------

# Single highlight colour to visually map commented text (light yellow)
_COMMENT_HIGHLIGHT = {"red": 1.00, "green": 0.95, "blue": 0.60}


def clear_old_highlights(doc_id: str, flags: list[dict]) -> None:
    """Reset background colour on all flagged ranges (remove old highlights)."""
    from googleapiclient.discovery import build

    creds = get_google_creds()
    docs = build("docs", "v1", credentials=creds, cache_discovery=False)

    requests = []
    for flag in flags:
        start = flag.get("start_index", 0)
        end = flag.get("end_index", 0)
        if start >= end:
            continue
        requests.append({
            "updateTextStyle": {
                "range": {"startIndex": start, "endIndex": end},
                "textStyle": {"backgroundColor": {"color": {"rgbColor": {"red": 1, "green": 1, "blue": 1}}}},
                "fields": "backgroundColor",
            }
        })

    if requests:
        docs.documents().batchUpdate(
            documentId=doc_id,
            body={"requests": requests},
        ).execute()


def highlight_flagged_paragraphs(doc_id: str, flags: list[dict]) -> int:
    """
    Add background colour highlights to flagged (non-compliant) paragraphs.
    Uses Google Docs API batchUpdate with UpdateTextStyle.
    Returns the number of paragraphs highlighted.
    """
    from googleapiclient.discovery import build

    creds = get_google_creds()
    docs = build("docs", "v1", credentials=creds, cache_discovery=False)

    requests = []
    for flag in flags:
        if flag["classification"] == "compliant":
            continue

        start = flag.get("start_index", 0)
        end = flag.get("end_index", 0)
        if start >= end:
            continue

        color = _COMMENT_HIGHLIGHT

        requests.append({
            "updateTextStyle": {
                "range": {"startIndex": start, "endIndex": end},
                "textStyle": {
                    "backgroundColor": {
                        "color": {"rgbColor": color},
                    },
                },
                "fields": "backgroundColor",
            }
        })

    if not requests:
        return 0

    # Google Docs batchUpdate has a limit; chunk if needed
    BATCH_SIZE = 50
    for chunk_start in range(0, len(requests), BATCH_SIZE):
        chunk = requests[chunk_start : chunk_start + BATCH_SIZE]
        docs.documents().batchUpdate(
            documentId=doc_id,
            body={"requests": chunk},
        ).execute()

    return len(requests)


# ---------------------------------------------------------------------------
# 9. BUILD OUTPUT
# ---------------------------------------------------------------------------

def build_flag(idx, inp, pb, sim, mt, rules, analysis) -> dict:
    return {
        "flag_id": f"FLAG_{idx:03d}",
        "input_clause_id": inp.id,
        "input_clause_section": inp.section,
        "input_text": inp.text,
        "matched_playbook_id": pb.id if mt != "new_clause" else None,
        "matched_playbook_text": pb.text if mt != "new_clause" else None,
        "similarity_score": round(sim, 4),
        "match_type": mt,
        "triggered_rules": [
            {"rule_id": r.rule_id, "source": r.source,
             "clause": r.clause, "risk": r.risk, "match_score": round(s, 4)}
            for r, s in rules
        ],
        "classification": analysis["classification"],
        "risk_level": analysis["risk_level"],
        "explanation": analysis["explanation"],
        "suggested_redline": analysis["suggested_redline"],
        "confidence": analysis.get("confidence", 0.5),
        "start_index": inp.start_index,
        "end_index": inp.end_index,
        "raw_text": inp.raw_text,
    }


def generate_summary(flags):
    by_cls = {}
    by_risk = {}
    for f in flags:
        by_cls[f["classification"]] = by_cls.get(f["classification"], 0) + 1
        by_risk[f["risk_level"]] = by_risk.get(f["risk_level"], 0) + 1

    ranked = sorted(flags, key=lambda x: (
        {"High": 0, "Medium": 1, "Low": 2}.get(x["risk_level"], 9),
        {"non_compliant": 0, "deviation_major": 1, "deviation_minor": 2, "compliant": 3}.get(x["classification"], 9),
    ))
    return {
        "total_clauses_analyzed": len(flags),
        "classification_breakdown": by_cls,
        "risk_breakdown": by_risk,
        "high_risk_count": sum(1 for f in flags if f["risk_level"] == "High"),
        "non_compliant_count": sum(1 for f in flags if f["classification"] == "non_compliant"),
        "top_risks": [
            {"flag_id": f["flag_id"], "section": f["input_clause_section"],
             "risk": f["risk_level"], "classification": f["classification"],
             "summary": f["explanation"][:200]}
            for f in ranked[:10]
        ],
    }


def generate_html_report(flags: list[dict], summary: dict, metadata: dict) -> str:
    """Generate a self-contained interactive HTML report."""
    import html as html_mod

    risk_colors = {"High": "#e74c3c", "Medium": "#f39c12", "Low": "#27ae60"}
    cls_colors = {
        "non_compliant": "#c0392b", "deviation_major": "#e74c3c",
        "deviation_minor": "#f39c12", "compliant": "#27ae60",
    }

    # Build table rows
    rows_html = ""
    for f in flags:
        risk_badge = f'<span class="badge" style="background:{risk_colors.get(f["risk_level"], "#999")}">{f["risk_level"]}</span>'
        cls_badge = f'<span class="badge" style="background:{cls_colors.get(f["classification"], "#999")}">{f["classification"].replace("_", " ").title()}</span>'
        confidence_pct = f'{f.get("confidence", 0) * 100:.0f}%'

        rules_html = ""
        if f["triggered_rules"]:
            rules_html = "<ul>" + "".join(
                f'<li>[{r["source"].upper()}] {html_mod.escape(r["clause"])} (Risk: {r["risk"]})</li>'
                for r in f["triggered_rules"]
            ) + "</ul>"

        pb_text = html_mod.escape(f["matched_playbook_text"] or "No playbook match")
        inp_text = html_mod.escape(f["input_text"])
        redline = html_mod.escape(f["suggested_redline"] or "None")
        explanation = html_mod.escape(f["explanation"])

        rows_html += f"""
        <tr class="flag-row" data-risk="{f['risk_level']}" data-cls="{f['classification']}">
            <td>{f['flag_id']}</td>
            <td>{html_mod.escape(f['input_clause_section'] or 'N/A')}</td>
            <td>{risk_badge}</td>
            <td>{cls_badge}</td>
            <td>{confidence_pct}</td>
            <td>{f['similarity_score']:.2f} ({f['match_type']})</td>
            <td><button class="expand-btn" onclick="toggleRow(this)">+</button></td>
        </tr>
        <tr class="detail-row" style="display:none">
            <td colspan="7">
                <div class="detail-grid">
                    <div class="detail-col">
                        <h4>Incoming Clause</h4>
                        <p class="clause-text">{inp_text}</p>
                    </div>
                    <div class="detail-col">
                        <h4>Playbook Clause</h4>
                        <p class="clause-text">{pb_text}</p>
                    </div>
                </div>
                <div class="detail-section">
                    <h4>Explanation</h4>
                    <p>{explanation}</p>
                </div>
                <div class="detail-section">
                    <h4>Triggered Rules</h4>
                    {rules_html or '<p>None</p>'}
                </div>
                <div class="detail-section">
                    <h4>Suggested Redline</h4>
                    <p class="redline">{redline}</p>
                </div>
            </td>
        </tr>"""

    risk_bd = summary.get("risk_breakdown", {})
    cls_bd = summary.get("classification_breakdown", {})

    report = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>DPA Contract Review Report</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f6fa; color: #2c3e50; padding: 20px; }}
  .container {{ max-width: 1200px; margin: 0 auto; }}
  h1 {{ font-size: 1.8em; margin-bottom: 5px; }}
  .subtitle {{ color: #7f8c8d; margin-bottom: 20px; }}
  .summary-cards {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 15px; margin-bottom: 25px; }}
  .card {{ background: white; border-radius: 10px; padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); text-align: center; }}
  .card .number {{ font-size: 2.2em; font-weight: 700; }}
  .card .label {{ color: #7f8c8d; font-size: 0.85em; margin-top: 5px; }}
  .charts {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 25px; }}
  .chart-box {{ background: white; border-radius: 10px; padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
  .chart-box canvas {{ max-height: 250px; }}
  .filters {{ margin-bottom: 15px; display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }}
  .filters select, .filters input {{ padding: 8px 12px; border: 1px solid #ddd; border-radius: 6px; font-size: 0.9em; }}
  table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
  th {{ background: #34495e; color: white; padding: 12px 15px; text-align: left; cursor: pointer; user-select: none; }}
  th:hover {{ background: #2c3e50; }}
  td {{ padding: 12px 15px; border-bottom: 1px solid #ecf0f1; }}
  .flag-row:hover {{ background: #f8f9fa; }}
  .badge {{ padding: 4px 10px; border-radius: 12px; color: white; font-size: 0.8em; font-weight: 600; }}
  .expand-btn {{ background: #3498db; color: white; border: none; border-radius: 50%; width: 28px; height: 28px; cursor: pointer; font-size: 1.1em; }}
  .expand-btn:hover {{ background: #2980b9; }}
  .detail-row td {{ background: #f8f9fa; padding: 20px; }}
  .detail-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 15px; }}
  .detail-col {{ background: white; border-radius: 8px; padding: 15px; border: 1px solid #e0e0e0; }}
  .detail-col h4 {{ color: #2c3e50; margin-bottom: 8px; }}
  .clause-text {{ font-size: 0.9em; line-height: 1.6; color: #555; }}
  .detail-section {{ margin-top: 12px; }}
  .detail-section h4 {{ color: #2c3e50; margin-bottom: 5px; }}
  .redline {{ background: #fff3cd; padding: 10px; border-radius: 6px; border-left: 3px solid #f39c12; font-style: italic; }}
  .print-btn {{ background: #2c3e50; color: white; border: none; padding: 10px 20px; border-radius: 6px; cursor: pointer; font-size: 0.9em; }}
  .print-btn:hover {{ background: #34495e; }}
  @media print {{ .filters, .print-btn, .expand-btn {{ display: none; }} .detail-row {{ display: table-row !important; }} }}
  @media (max-width: 768px) {{ .charts, .detail-grid {{ grid-template-columns: 1fr; }} }}
</style>
</head>
<body>
<div class="container">
  <h1>DPA Contract Review Report</h1>
  <p class="subtitle">Generated by {html_mod.escape(metadata.get('tool', 'DPA Review Tool'))} | Mode: {html_mod.escape(metadata.get('analysis_mode', 'N/A'))} | Model: {html_mod.escape(metadata.get('embedding_model', 'N/A'))}</p>

  <div class="summary-cards">
    <div class="card"><div class="number">{summary['total_clauses_analyzed']}</div><div class="label">Clauses Analyzed</div></div>
    <div class="card"><div class="number" style="color:#e74c3c">{summary['high_risk_count']}</div><div class="label">High Risk</div></div>
    <div class="card"><div class="number" style="color:#c0392b">{summary['non_compliant_count']}</div><div class="label">Non-Compliant</div></div>
    <div class="card"><div class="number" style="color:#27ae60">{cls_bd.get('compliant', 0)}</div><div class="label">Compliant</div></div>
  </div>

  <div class="charts">
    <div class="chart-box"><canvas id="riskChart"></canvas></div>
    <div class="chart-box"><canvas id="clsChart"></canvas></div>
  </div>

  <div class="filters">
    <select id="filterRisk" onchange="filterTable()">
      <option value="">All Risks</option>
      <option value="High">High</option>
      <option value="Medium">Medium</option>
      <option value="Low">Low</option>
    </select>
    <select id="filterCls" onchange="filterTable()">
      <option value="">All Classifications</option>
      <option value="compliant">Compliant</option>
      <option value="deviation_minor">Deviation Minor</option>
      <option value="deviation_major">Deviation Major</option>
      <option value="non_compliant">Non-Compliant</option>
    </select>
    <input type="text" id="filterSearch" placeholder="Search clauses..." oninput="filterTable()">
    <button class="print-btn" onclick="window.print()">Export PDF</button>
  </div>

  <table>
    <thead>
      <tr>
        <th onclick="sortTable(0)">Flag ID</th>
        <th onclick="sortTable(1)">Section</th>
        <th onclick="sortTable(2)">Risk</th>
        <th onclick="sortTable(3)">Classification</th>
        <th onclick="sortTable(4)">Confidence</th>
        <th onclick="sortTable(5)">Similarity</th>
        <th>Details</th>
      </tr>
    </thead>
    <tbody id="flagsBody">
      {rows_html}
    </tbody>
  </table>
</div>

<script>
function toggleRow(btn) {{
  const detailRow = btn.closest('tr').nextElementSibling;
  const isHidden = detailRow.style.display === 'none';
  detailRow.style.display = isHidden ? 'table-row' : 'none';
  btn.textContent = isHidden ? '-' : '+';
}}

function filterTable() {{
  const risk = document.getElementById('filterRisk').value;
  const cls = document.getElementById('filterCls').value;
  const search = document.getElementById('filterSearch').value.toLowerCase();
  const rows = document.querySelectorAll('.flag-row');
  rows.forEach(row => {{
    const detail = row.nextElementSibling;
    const matchRisk = !risk || row.dataset.risk === risk;
    const matchCls = !cls || row.dataset.cls === cls;
    const matchSearch = !search || row.textContent.toLowerCase().includes(search) ||
                        detail.textContent.toLowerCase().includes(search);
    const show = matchRisk && matchCls && matchSearch;
    row.style.display = show ? '' : 'none';
    if (!show) detail.style.display = 'none';
  }});
}}

let sortDir = {{}};
function sortTable(col) {{
  const tbody = document.getElementById('flagsBody');
  const flagRows = Array.from(tbody.querySelectorAll('.flag-row'));
  sortDir[col] = !sortDir[col];
  flagRows.sort((a, b) => {{
    const aVal = a.children[col].textContent.trim();
    const bVal = b.children[col].textContent.trim();
    const aNum = parseFloat(aVal);
    const bNum = parseFloat(bVal);
    let cmp = (!isNaN(aNum) && !isNaN(bNum)) ? aNum - bNum : aVal.localeCompare(bVal);
    return sortDir[col] ? cmp : -cmp;
  }});
  flagRows.forEach(row => {{
    const detail = row.nextElementSibling;
    tbody.appendChild(row);
    tbody.appendChild(detail);
  }});
}}

// Charts
const riskCtx = document.getElementById('riskChart').getContext('2d');
new Chart(riskCtx, {{
  type: 'doughnut',
  data: {{
    labels: {json.dumps(list(risk_bd.keys()))},
    datasets: [{{ data: {json.dumps(list(risk_bd.values()))},
      backgroundColor: {json.dumps([risk_colors.get(k, '#999') for k in risk_bd.keys()])} }}]
  }},
  options: {{ responsive: true, plugins: {{ title: {{ display: true, text: 'Risk Breakdown' }} }} }}
}});

const clsCtx = document.getElementById('clsChart').getContext('2d');
new Chart(clsCtx, {{
  type: 'doughnut',
  data: {{
    labels: {json.dumps([k.replace('_', ' ').title() for k in cls_bd.keys()])},
    datasets: [{{ data: {json.dumps(list(cls_bd.values()))},
      backgroundColor: {json.dumps([cls_colors.get(k, '#999') for k in cls_bd.keys()])} }}]
  }},
  options: {{ responsive: true, plugins: {{ title: {{ display: true, text: 'Classification Breakdown' }} }} }}
}});
</script>
</body>
</html>"""

    report_path = OUTPUT_DIR / "report.html"
    with open(report_path, "w") as f:
        f.write(report)
    return str(report_path)


def print_rich_summary(summary: dict, flags: list[dict], metadata: dict) -> None:
    """Print a rich formatted summary to the terminal."""
    try:
        from rich.console import Console
        from rich.table import Table
        from rich.panel import Panel
        from rich.columns import Columns
        from rich import box
    except ImportError:
        # Fallback to plain text if rich not installed
        return _print_plain_summary(summary, flags)

    console = Console()
    console.print()

    # Executive summary panel
    risk_bd = summary.get("risk_breakdown", {})
    cls_bd = summary.get("classification_breakdown", {})
    summary_text = (
        f"[bold]Clauses Analyzed:[/bold] {summary['total_clauses_analyzed']}\n"
        f"[bold red]High Risk:[/bold red] {risk_bd.get('High', 0)}  "
        f"[bold yellow]Medium:[/bold yellow] {risk_bd.get('Medium', 0)}  "
        f"[bold green]Low:[/bold green] {risk_bd.get('Low', 0)}\n"
        f"[bold]Compliant:[/bold] {cls_bd.get('compliant', 0)}  "
        f"[bold yellow]Minor Dev:[/bold yellow] {cls_bd.get('deviation_minor', 0)}  "
        f"[bold red]Major Dev:[/bold red] {cls_bd.get('deviation_major', 0)}  "
        f"[bold]Non-Compliant:[/bold] {cls_bd.get('non_compliant', 0)}\n"
        f"[bold]Mode:[/bold] {metadata.get('analysis_mode', 'N/A')}  "
        f"[bold]Model:[/bold] {metadata.get('embedding_model', 'N/A')}"
    )
    console.print(Panel(summary_text, title="DPA Review Summary", border_style="blue", expand=False))

    # Top risks table
    table = Table(title="Top Risk Flags", box=box.ROUNDED, show_lines=True)
    table.add_column("Flag", style="bold", width=10)
    table.add_column("Section", width=25)
    table.add_column("Risk", width=8)
    table.add_column("Classification", width=18)
    table.add_column("Confidence", width=10)
    table.add_column("Summary", width=60)

    risk_style = {"High": "bold red", "Medium": "bold yellow", "Low": "bold green"}

    ranked = sorted(flags, key=lambda x: (
        {"High": 0, "Medium": 1, "Low": 2}.get(x["risk_level"], 9),
        {"non_compliant": 0, "deviation_major": 1, "deviation_minor": 2, "compliant": 3}.get(x["classification"], 9),
    ))

    for f in ranked[:10]:
        if f["classification"] == "compliant":
            continue
        table.add_row(
            f["flag_id"],
            f["input_clause_section"] or "N/A",
            f"[{risk_style.get(f['risk_level'], '')}]{f['risk_level']}[/]",
            f["classification"].replace("_", " ").title(),
            f'{f.get("confidence", 0) * 100:.0f}%',
            f["explanation"][:80] + "..." if len(f["explanation"]) > 80 else f["explanation"],
        )

    console.print(table)
    console.print()


def _print_plain_summary(summary: dict, flags: list[dict]) -> None:
    """Fallback plain-text summary when rich is not available."""
    print(f"\n{'='*60}")
    print(f"  SUMMARY")
    print(f"{'='*60}")
    print(f"  Clauses analysed : {summary['total_clauses_analyzed']}")
    print(f"  Classification   : {summary['classification_breakdown']}")
    print(f"  Risk breakdown   : {summary['risk_breakdown']}")
    print(f"  High-risk flags  : {summary['high_risk_count']}")
    print(f"  Non-compliant    : {summary['non_compliant_count']}")
    if summary["top_risks"]:
        print(f"\n  TOP RISKS:")
        for r in summary["top_risks"][:5]:
            print(f"    [{r['risk']:6s}] {r['flag_id']} "
                  f"({r['section'] or 'N/A'}): {r['classification']}")
    print()


# ---------------------------------------------------------------------------
# 10. MAIN PIPELINE
# ---------------------------------------------------------------------------

def main() -> None:
    # ---- Parse args ----
    args = sys.argv[1:]
    if not args:
        print("Usage: python main.py <input_doc_url_or_id_or_file> [--playbook <url_or_id>] [--mode heuristic|llm|hybrid|full]")
        print("\nExamples:")
        print("  python main.py input.docx --mode hybrid         # Default: heuristic + LLM for flagged clauses")
        print("  python main.py input.docx --mode full            # Full-context LLM, para by para")
        print("  python main.py https://docs.google.com/document/d/ABC123/edit --mode full")
        print("  python main.py ABC123_DOC_ID --playbook XYZ789_DOC_ID --mode llm")
        sys.exit(0)

    playbook_gdoc = None
    input_arg = None
    analysis_mode = "hybrid"  # default: heuristic + LLM for flagged clauses
    i = 0
    while i < len(args):
        if args[i] == "--playbook" and i + 1 < len(args):
            playbook_gdoc = args[i + 1]
            i += 2
        elif args[i] == "--mode" and i + 1 < len(args):
            analysis_mode = args[i + 1].lower()
            if analysis_mode not in ("heuristic", "llm", "hybrid", "full"):
                print(f"Error: --mode must be heuristic, llm, hybrid, or full (got '{analysis_mode}')")
                sys.exit(1)
            i += 2
        else:
            input_arg = args[i]
            i += 1

    if not input_arg:
        input_arg = "input.docx"

    # Determine if input is a local file or Google Doc
    input_is_local = Path(input_arg).suffix in (".docx", ".doc") or (
        BASE_DIR / input_arg
    ).exists()
    input_doc_id = None

    if input_is_local:
        input_path = Path(input_arg)
        if not input_path.is_absolute():
            input_path = BASE_DIR / input_path
        if not input_path.exists():
            print(f"Error: File not found: {input_path}")
            sys.exit(1)
    else:
        input_doc_id = extract_doc_id(input_arg)

    # LLM availability
    has_llm = bool(ANTHROPIC_API_KEY)
    if analysis_mode in ("llm", "hybrid", "full") and not has_llm:
        print(f"Warning: --mode {analysis_mode} requested but ANTHROPIC_API_KEY not set. Falling back to heuristic.")
        analysis_mode = "heuristic"

    # Full-context mode: delegate to the new pipeline
    if analysis_mode == "full":
        print("DPA Contract Review Tool")
        print(f"Analysis mode: Full-Context LLM (paragraph by paragraph)")
        print(f"Input: {input_doc_id or input_path}")
        print()
        from contract_review.pipeline import run_pipeline
        result = run_pipeline(
            input_source=input_arg,
            reviewer="",
        )
        return

    mode_display = {
        "heuristic": "Heuristic only",
        "llm": f"LLM ({LLM_MODEL})",
        "hybrid": f"Hybrid (heuristic + LLM for flagged clauses)",
    }[analysis_mode]

    print("DPA Contract Review Tool")
    print(f"Analysis mode: {mode_display}")
    if input_doc_id:
        print(f"Input: Google Doc {input_doc_id}")
    else:
        print(f"Input: {input_path}")
    print()

    # ---- Step 1: Load rulebook from xlsx ----
    print("[Step 1/6] Loading rulebook from xlsx...")
    rules = load_rulebook(RULEBOOK_PATH)
    print(f"  Loaded {len(rules)} rules ({sum(1 for r in rules if r.source == 'legal')} legal, "
          f"{sum(1 for r in rules if r.source == 'infosec')} infosec)")

    # ---- Step 2: Fetch paragraphs ----
    print("\n[Step 2/6] Fetching paragraphs...")
    if input_doc_id:
        input_paras = fetch_gdoc_paragraphs(input_doc_id)
    else:
        input_paras = fetch_docx_paragraphs(input_path)
    print(f"  Input: {len(input_paras)} paragraphs")

    if playbook_gdoc:
        pb_id = extract_doc_id(playbook_gdoc)
        pb_paras = fetch_gdoc_paragraphs(pb_id)
    else:
        pb_paras = fetch_docx_paragraphs(PLAYBOOK_PATH)
    print(f"  Playbook: {len(pb_paras)} paragraphs")

    # ---- Step 3: Extract clauses ----
    print("\n[Step 3/6] Extracting clauses...")
    input_clauses = extract_clauses(input_paras, "input")
    playbook_clauses = extract_clauses(pb_paras, "playbook")
    print(f"  Input clauses:    {len(input_clauses)}")
    print(f"  Playbook clauses: {len(playbook_clauses)}")

    if not input_clauses:
        print("Error: No clauses extracted from input.")
        sys.exit(1)
    if not playbook_clauses:
        print("Error: No clauses extracted from playbook.")
        sys.exit(1)

    # ---- Step 4: Match clauses + rules ----
    print("\n[Step 4/6] Matching clauses to playbook + rules...")
    matches = match_clauses(input_clauses, playbook_clauses)
    strong = sum(1 for *_, mt in matches if mt == "strong")
    partial = sum(1 for *_, mt in matches if mt == "partial")
    new = sum(1 for *_, mt in matches if mt == "new_clause")
    print(f"  Strong: {strong}  |  Partial: {partial}  |  New: {new}")

    clause_rules_raw = [match_rules(inp, rules) for inp, _, _, _ in matches]
    # Apply rule specificity scoring to penalize overly broad rules
    clause_rules = apply_rule_specificity(clause_rules_raw)
    triggered = sum(1 for cr in clause_rules if cr)
    print(f"  Clauses with triggered rules: {triggered}/{len(matches)}")

    # ---- Step 5: Analyse each clause ----
    print("\n[Step 5/6] Analysing clauses...")
    flags: list[dict] = []
    total = len(matches)
    llm_calls = 0
    for n, ((inp, pb, sim, mt), crules) in enumerate(
        zip(matches, clause_rules), start=1,
    ):
        label = inp.section or inp.id

        if analysis_mode == "heuristic":
            # Pure heuristic mode — never call LLM
            if mt == "strong" and not crules:
                analysis = dict(
                    classification="compliant", risk_level="Low",
                    explanation=f"Matches ClearTax standard (sim={sim:.2f}).",
                    suggested_redline="", confidence=_compute_confidence(sim, crules),
                )
            else:
                print(f"  [{n}/{total}] {label}...")
                analysis = _heuristic(inp, pb, sim, mt, crules)

        elif analysis_mode == "llm":
            # Pure LLM mode — always call LLM
            print(f"  [{n}/{total}] {label} (LLM)...")
            analysis = analyze_clause(inp, pb, sim, mt, crules, True)
            llm_calls += 1

        else:
            # Hybrid mode — heuristic first, LLM only for non-compliant/uncertain
            heuristic_result = _heuristic(inp, pb, sim, mt, crules)
            if heuristic_result["classification"] == "compliant" and not crules:
                # Fast path: clearly compliant, skip LLM
                analysis = heuristic_result
            else:
                # Non-compliant or uncertain → send to LLM for quality explanation
                print(f"  [{n}/{total}] {label} (LLM)...")
                analysis = analyze_clause(inp, pb, sim, mt, crules, True)
                llm_calls += 1

        flags.append(build_flag(n, inp, pb, sim, mt, crules, analysis))

    if analysis_mode != "heuristic":
        print(f"  LLM calls made: {llm_calls}/{total} clauses")

    # ---- Step 6: Output ----
    print("\n[Step 6/6] Generating output...")
    OUTPUT_DIR.mkdir(exist_ok=True)
    summary = generate_summary(flags)

    output_metadata = {
        "tool": "DPA Contract Review Tool",
        "input_source": input_doc_id or str(input_path.name),
        "playbook_source": playbook_gdoc or PLAYBOOK_PATH.name,
        "rulebook": RULEBOOK_PATH.name,
        "rules_loaded": len(rules),
        "analysis_mode": analysis_mode,
        "embedding_model": EMBED_MODEL,
    }
    output = {
        "metadata": output_metadata,
        "summary": summary,
        "flags": flags,
    }
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"  flags.json written to: {OUTPUT_PATH}")

    # Generate HTML report
    report_path = generate_html_report(flags, summary, output_metadata)
    print(f"  HTML report written to: {report_path}")

    # Add comments to Google Doc — only where there are real issues
    if input_doc_id:
        issue_flags = [f for f in flags if f["classification"] != "compliant"]
        print(f"\n  Clearing old review comments...")
        deleted = clear_old_comments(input_doc_id)
        if deleted:
            print(f"  Removed {deleted} old comments.")
        # Clear old highlights, then re-apply for flagged paragraphs
        print(f"  Clearing old highlights...")
        clear_old_highlights(input_doc_id, flags)

        print(f"  Adding comments to {len(issue_flags)} flagged paragraphs...")
        added = add_comments_to_doc(input_doc_id, flags)
        print(f"  {added} comments added to Google Doc.")

        print(f"  Highlighting flagged paragraphs...")
        highlighted = highlight_flagged_paragraphs(input_doc_id, flags)
        print(f"  {highlighted} paragraphs highlighted (yellow = has comment in sidebar).")

    # ---- Print summary (rich or plain) ----
    print_rich_summary(summary, flags, output_metadata)


if __name__ == "__main__":
    main()
