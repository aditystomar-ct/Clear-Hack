"""Clause extraction, rulebook loading, paragraph fetching."""

import re
import sys
from pathlib import Path

from .models import Clause, Rule
from .auth import get_google_creds


# ---------------------------------------------------------------------------
# Extract Google Doc ID
# ---------------------------------------------------------------------------

def extract_doc_id(url_or_id: str) -> str:
    """Extract Google Doc ID from a URL or return as-is if already an ID."""
    m = re.search(r"/document/d/([a-zA-Z0-9_-]+)", url_or_id)
    if m:
        return m.group(1)
    if re.match(r"^[a-zA-Z0-9_-]{20,}$", url_or_id):
        return url_or_id
    print(f"Error: Cannot parse Google Doc ID from: {url_or_id}")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Load Rulebook from XLSX
# ---------------------------------------------------------------------------

def load_rulebook(path: Path) -> list[Rule]:
    """
    Parse DPA Rulebook.xlsx dynamically.
    Reads Legal and Infosec sheets, extracts rules with clause/subclause/risk/response.
    """
    import openpyxl

    if not path.exists():
        print(f"Error: Rulebook not found: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(path), read_only=True)
    rules: list[Rule] = []
    rule_idx = 0

    for sheet_name in wb.sheetnames:
        name_lower = sheet_name.lower()
        if "legal" in name_lower:
            source = "legal"
        elif "infosec" in name_lower:
            source = "infosec"
        else:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        header = [str(c).lower().strip() if c else "" for c in rows[0]]

        def find_col(*keywords):
            for i, h in enumerate(header):
                if any(kw in h for kw in keywords):
                    return i
            return None

        col_clause = find_col("clause")
        col_sub = find_col("sub-clause", "subclause", "sub clause")
        col_risk = find_col("risk")
        col_resp = find_col("response")

        if col_clause is None and col_sub is None:
            continue

        if col_sub is None:
            col_sub = col_clause

        current_clause = ""
        for row in rows[1:]:
            def cell(idx):
                if idx is None or idx >= len(row):
                    return ""
                return str(row[idx]).strip() if row[idx] else ""

            clause_val = cell(col_clause)
            sub_val = cell(col_sub)
            risk_val = cell(col_risk)
            resp_val = cell(col_resp)

            if clause_val:
                current_clause = clause_val

            if not risk_val or risk_val.lower() in ("none", "risk", ""):
                continue
            if not resp_val:
                continue

            rule_idx += 1
            rules.append(Rule(
                rule_id=f"{source}_{rule_idx}",
                source=source,
                clause=current_clause,
                subclause=sub_val or current_clause,
                risk=risk_val,
                response=resp_val,
            ))

    wb.close()
    return rules


# ---------------------------------------------------------------------------
# Fetch Paragraphs
# ---------------------------------------------------------------------------

def fetch_gdoc_paragraphs(doc_id: str) -> list[dict]:
    """Fetch paragraphs from a Google Doc via the Docs API."""
    import requests as _requests
    from google.auth.transport.requests import AuthorizedSession

    creds = get_google_creds()
    session = AuthorizedSession(creds)
    session.timeout = 60

    resp = session.get(f"https://docs.googleapis.com/v1/documents/{doc_id}")
    resp.raise_for_status()
    doc = resp.json()

    paragraphs: list[dict] = []
    for element in doc.get("body", {}).get("content", []):
        para = element.get("paragraph")
        if not para:
            continue
        start = element.get("startIndex", 0)
        end = element.get("endIndex", 0)
        parts = []
        for elem in para.get("elements", []):
            tr = elem.get("textRun")
            if tr:
                parts.append(tr.get("content", ""))
        full = "".join(parts).strip()
        if full:
            paragraphs.append({
                "text": full,
                "start_index": start,
                "end_index": end,
            })
    return paragraphs


def fetch_docx_paragraphs(path: Path) -> list[dict]:
    """Read paragraphs from a local .docx file."""
    from docx import Document

    doc = Document(str(path))
    paragraphs: list[dict] = []
    offset = 0
    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            paragraphs.append({
                "text": text,
                "start_index": offset,
                "end_index": offset + len(text),
            })
            offset += len(text) + 1
    return paragraphs


# ---------------------------------------------------------------------------
# Clause Extraction (content-based filtering)
# ---------------------------------------------------------------------------

_CLAUSE_VERBS = {
    "shall", "will", "must", "may", "agrees", "acknowledges",
    "warrants", "represents", "confirms", "ensures", "undertakes",
}

_DEF_RE = re.compile(
    r'^\s*[\"\u201c\u2018][^\"\u201d\u2019]+[\"\u201d\u2019]\s*'
    r'(,\s*[\"\u201c\u2018][^\"\u201d\u2019]+[\"\u201d\u2019]\s*)*'
    r'(means|shall\s+mean|shall\s+have)',
    re.IGNORECASE,
)

_SKIP_RE = re.compile(
    r"(^IN WITNESS WHEREOF|^WHEREAS\b|^NOW,?\s*THEREFORE|"
    r"^Sign\s*:|^Signed?\s*:|_{5,})",
    re.IGNORECASE,
)

_APPENDIX_RE = re.compile(
    r"^(APPENDIX|ANNEX|ATTACHMENT)\s", re.IGNORECASE,
)


def _is_definition(text: str) -> bool:
    low = text.lower()
    if "shall mean" in low or "shall have the meaning" in low:
        return True
    return bool(_DEF_RE.match(text))


def _is_section_header(text: str) -> tuple[bool, str, str]:
    stripped = text.strip()
    if len(stripped) < 100 and stripped.upper() == stripped and stripped.isascii():
        if any(c.isalpha() for c in stripped):
            return True, stripped.rstrip(".:;, "), ""
    m = re.match(
        r'^(\d{1,3}\.?\s+)?([A-Z][A-Za-z,;/&\s\-\(\)\']+?)\.\s*(.*)',
        stripped, re.DOTALL,
    )
    if m:
        label = m.group(2).strip()
        remainder = m.group(3).strip()
        if len(label) > 100:
            return False, "", ""
        if _CLAUSE_VERBS & set(label.lower().split()):
            return False, "", ""
        if len(label.split()) <= 15:
            return True, label, remainder
    return False, "", ""


def _is_appendix_boundary(text: str) -> bool:
    stripped = text.strip()
    up = stripped.upper()
    if up in ("EXHIBIT A", "EXHIBIT B", "EXHIBIT C"):
        return False
    if _APPENDIX_RE.match(stripped):
        return True
    if re.match(
        r"^(STANDARD CONTRACTUAL|LIST OF SUB-?PROCESSORS|"
        r"TECHNICAL AND ORGANI[SZ]ATIONAL|DESCRIPTION OF TECHNICAL)",
        stripped, re.IGNORECASE,
    ):
        return True
    return False


def _is_preamble(text: str) -> bool:
    low = text.lower()
    return any(p in low for p in [
        "this data processing agreement",
        "the dpa shall form",
        "entering into this dpa",
        "in the event of inconsistencies",
        "the parties have agreed",
        "order of priority shall be",
        "hereinafter referred to",
        "referred to individually as",
        "seek to implement a data processing",
        "wish to lay down their rights",
        "in consideration of the mutual covenants",
    ])


def extract_clauses(paragraphs: list[dict], source: str) -> list[Clause]:
    """Extract substantive clauses from a list of paragraphs."""
    clauses: list[Clause] = []
    section = ""
    in_appendix = False
    in_definitions = False
    idx = 0
    i = 0

    while i < len(paragraphs):
        p = paragraphs[i]
        text = p["text"]

        if _is_appendix_boundary(text):
            in_appendix = True
            i += 1
            continue
        if in_appendix:
            i += 1
            continue

        if _SKIP_RE.search(text):
            i += 1
            continue

        is_hdr, sec_name, remainder = _is_section_header(text)
        if is_hdr:
            section = sec_name
            in_definitions = any(
                kw in sec_name.lower()
                for kw in ("definition", "defined term", "interpretation")
            )
            if len(remainder) >= 40:
                text = remainder
            else:
                i += 1
                continue

        if in_definitions:
            i += 1
            continue
        if _is_definition(text):
            i += 1
            continue

        if _is_preamble(text):
            i += 1
            continue

        if len(text) < 35:
            i += 1
            continue

        merged_text = text
        merged_end = p["end_index"]
        while i + 1 < len(paragraphs):
            nxt = paragraphs[i + 1]["text"]
            if len(nxt) >= 40:
                break
            if _SKIP_RE.search(nxt) or _is_appendix_boundary(nxt) or _is_definition(nxt):
                break
            nxt_hdr, _, _ = _is_section_header(nxt)
            if nxt_hdr:
                break
            merged_text += " " + nxt
            merged_end = paragraphs[i + 1]["end_index"]
            i += 1

        if len(merged_text) >= 40:
            idx += 1
            clauses.append(Clause(
                id=f"{source}_{idx}",
                text=merged_text,
                section=section,
                source=source,
                start_index=p["start_index"],
                end_index=merged_end,
                raw_text=p["text"],
            ))

        i += 1

    return clauses
